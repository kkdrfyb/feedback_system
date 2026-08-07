"""Spreadsheets Router — 在线 Excel 表格协作"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, timezone
import json, io

from openpyxl import load_workbook, Workbook
import re

_EXCEL_TITLE_ILLEGAL = re.compile(r"[\[\]\:\*\?\/\\]")

try:
    from .. import models, schemas
    from ..database import get_db
    from ..auth import get_current_user
except (ImportError, ValueError):
    import models, schemas
    from database import get_db
    from auth import get_current_user

router = APIRouter(tags=["spreadsheets"], dependencies=[Depends(get_current_user)])

# ── helpers ──────────────────────────────────────────────────────────────────

def _can_edit_row(row: models.SpreadsheetRow, spreadsheet: models.Spreadsheet, current_user: models.User) -> bool:
    """行级权限：管理员/创建者 或 该行负责人 可编辑；无负责人列时全员可编辑"""
    if current_user.role == "admin" or spreadsheet.creator_id == current_user.id:
        return True
    if spreadsheet.owner_column:
        try:
            data = json.loads(row.data or "{}")
        except json.JSONDecodeError:
            data = {}
        owner_username = (data.get(spreadsheet.owner_column) or "").strip().lstrip("@")
        if owner_username == current_user.username:
            return True
        return False  # 有负责人列但用户不是该行负责人 → 禁止编辑
    return True  # 无负责人列时全员可编辑


def _row_to_dict(row: models.SpreadsheetRow) -> dict:
    try:
        data = json.loads(row.data or "{}")
    except json.JSONDecodeError:
        data = {}
    return {
        "id": row.id,
        "row_index": row.row_index,
        "data": data,
        "last_edited_by": row.last_edited_by,
        "last_edited_at": row.last_edited_at.isoformat() if row.last_edited_at else None,
    }


# ── routes ───────────────────────────────────────────────────────────────────

@router.post("/spreadsheets/upload")
async def upload_spreadsheet(
    file: UploadFile = File(...),
    title: str = Form(""),
    description: str = Form(""),
    owner_column: str = Form(""),  # 标识负责人列的列名（可选）
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只支持 .xlsx 格式的 Excel 文件")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10 MB 上限
        raise HTTPException(status_code=400, detail="文件大小不能超过 10 MB")
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件")
    ws = wb.active
    rows_data = list(ws.iter_rows(values_only=True, max_row=200))  # 最多 200 行
    if len(rows_data) < 2:
        raise HTTPException(status_code=400, detail="Excel 至少需要表头 + 1 行数据")

    # 第一行是表头
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows_data[0])]
    columns = [{"key": f"c{i}", "label": h, "width": 180 if i == 0 else 150} for i, h in enumerate(headers)]
    owner_key = ""
    if owner_column:
        for col in columns:
            if col["label"] == owner_column:
                owner_key = col["key"]
                break

    title = title or (file.filename.rsplit(".", 1)[0])
    sheet = models.Spreadsheet(
        title=title,
        description=description or "",
        creator_id=current_user.id,
        columns=json.dumps(columns, ensure_ascii=False),
        owner_column=owner_key or None,
    )
    db.add(sheet)
    db.flush()

    # 数据行（第2行开始）
    for idx, row_vals in enumerate(rows_data[1:]):
        data = {f"c{i}": (str(v) if v is not None else "") for i, v in enumerate(row_vals)}
        db.add(models.SpreadsheetRow(
            spreadsheet_id=sheet.id,
            row_index=idx,
            data=json.dumps(data, ensure_ascii=False),
        ))
    db.commit()
    db.refresh(sheet)
    return {"id": sheet.id, "title": title, "row_count": len(rows_data) - 1}


@router.get("/spreadsheets")
def list_spreadsheets(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    sheets = db.query(models.Spreadsheet).order_by(
        models.Spreadsheet.created_at.desc()
    ).all()
    result = []
    for s in sheets:
        row_count = db.query(models.SpreadsheetRow).filter(
            models.SpreadsheetRow.spreadsheet_id == s.id
        ).count()
        result.append({
            "id": s.id,
            "title": s.title,
            "creator_id": s.creator_id,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "row_count": row_count,
        })
    return result


@router.get("/spreadsheets/{sheet_id}")
def get_spreadsheet(
    sheet_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    sheet = db.query(models.Spreadsheet).filter(models.Spreadsheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="表格不存在")
    rows = db.query(models.SpreadsheetRow).filter(
        models.SpreadsheetRow.spreadsheet_id == sheet_id
    ).order_by(models.SpreadsheetRow.row_index).all()

    try:
        cols = json.loads(sheet.columns) if sheet.columns else []
    except json.JSONDecodeError:
        cols = []

    return {
        "id": sheet.id,
        "title": sheet.title,
        "description": sheet.description,
        "creator_id": sheet.creator_id,
        "columns": cols,
        "owner_column": sheet.owner_column,
        "created_at": sheet.created_at.isoformat() if sheet.created_at else None,
        "rows": [_row_to_dict(r) for r in rows],
    }


@router.put("/spreadsheets/{sheet_id}/cells")
def update_cell(
    sheet_id: int,
    payload: schemas.CellUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    sheet = db.query(models.Spreadsheet).filter(models.Spreadsheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="表格不存在")
    row = db.query(models.SpreadsheetRow).filter(
        models.SpreadsheetRow.id == payload.row_id,
        models.SpreadsheetRow.spreadsheet_id == sheet_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="行不存在")
    if not _can_edit_row(row, sheet, current_user):
        raise HTTPException(status_code=403, detail="你没有权限编辑这一行")

    try:
        data = json.loads(row.data or "{}")
    except json.JSONDecodeError:
        data = {}
    data[payload.key] = (payload.value or "")
    row.data = json.dumps(data, ensure_ascii=False)
    row.last_edited_by = current_user.id
    row.last_edited_at = datetime.now(timezone.utc)
    sheet.updated_at = datetime.now(timezone.utc)

    db.commit()
    return {"row_id": row.id, "key": payload.key, "value": payload.value}


@router.get("/spreadsheets/{sheet_id}/export")
def export_spreadsheet(
    sheet_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    from fastapi.responses import StreamingResponse

    sheet = db.query(models.Spreadsheet).filter(models.Spreadsheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="表格不存在")
    rows = db.query(models.SpreadsheetRow).filter(
        models.SpreadsheetRow.spreadsheet_id == sheet_id
    ).order_by(models.SpreadsheetRow.row_index).all()

    try:
        cols = json.loads(sheet.columns) if sheet.columns else []
    except json.JSONDecodeError:
        cols = []

    wb = Workbook()
    ws = wb.active
    safe_title = _EXCEL_TITLE_ILLEGAL.sub("_", sheet.title)[:31]
    ws.title = safe_title or "Sheet1"
    # 表头
    for ci, col in enumerate(cols, 1):
        ws.cell(row=1, column=ci, value=col["label"])
    # 数据
    for ri, row in enumerate(rows, 2):
        try:
            data = json.loads(row.data or "{}")
        except json.JSONDecodeError:
            data = {}
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=data.get(col["key"], ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=spreadsheet_{sheet_id}_export.xlsx"},
    )


@router.delete("/spreadsheets/{sheet_id}")
def delete_spreadsheet(
    sheet_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    sheet = db.query(models.Spreadsheet).filter(models.Spreadsheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="表格不存在")
    if current_user.role != "admin" and sheet.creator_id != current_user.id:
        raise HTTPException(status_code=403, detail="只有创建者或管理员可以删除")
    db.query(models.SpreadsheetRow).filter(
        models.SpreadsheetRow.spreadsheet_id == sheet_id
    ).delete()
    db.delete(sheet)
    db.commit()
    return {"detail": "表格已删除"}
